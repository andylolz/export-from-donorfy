from openpyxl import Workbook

import donorfy as df

tables = [
    ('Activities', '1955351a-2672-40fe-8097-7a57ca1adbd8'),
    ('Connections', 'e196d7c4-903b-437f-91df-bef643474edf'),
    ('Constituents', '4c6f6a32-517d-4d39-bc65-cf642075977b'),
    ('Gift Aid', '2535e0ea-34d1-4bd5-a9ad-79512f32d1e3'),
    ('Opportunities', 'f8a8dc02-9b5d-4559-a297-2354cb6703bf'),
    ('Opportunity Pledges', 'ea5d8d21-3ac7-44a7-b398-0c4057b71eb4'),
    ('Recurring Payment Instructions', '6ed9931b-b94b-455c-9b84-ee1aa3571d19'),
    ('Soft Credits', '27af90c1-971a-4029-a20f-dc4746516c1c'),
    ('Tags', 'a29b2121-8d90-4b61-b72b-0e93a72779b4'),
    ('Transactions', '7d897c73-11da-4c49-8d46-6595f87a78b2'),
]

wb = Workbook()
del wb['Sheet']

for list_name, list_id in tables:
    print('Fetching {} ...'.format(list_name))
    ws = wb.create_sheet(title=list_name)
    for row, member in enumerate(df.get_list_members(list_id)):
        if row == 0:
            for col, header in enumerate(member.keys()):
                ws.cell(1, col + 1, header)
        for col, val in enumerate(member.values()):
            ws.cell(row + 2, col + 1, val)

wb.save(filename='output.xlsx')
